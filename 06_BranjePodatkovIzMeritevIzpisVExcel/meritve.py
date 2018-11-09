#####################################################################
################## GET METADATA FROM XLS FILES##  #################
################## AND SAVE THEM TO EXCEL TABLE ###################
#####################################################################

import openpyxl
import pyperclip
import os,sys
from openpyxl.workbook import Workbook
from lxml import etree
import pyautogui
from openpyxl.styles import PatternFill
import random

##################BRANJE XML DATOTEK####################################################
# file = '7060_2018_07_13.mmt'
# et = etree.parse(file)
# ProcessedBy = et.xpath("Project/Site_Information/ProcessedBy/text()")[0]
# Measurement_Date = et.xpath("Project/Site_Information/Measurement_Date/text()")[0]
# River_Name = et.xpath("Project/Site_Information/River_Name/text()")[0]
# Name = et.xpath("Project/Site_Information/Name/text()")[0]
# Remarks = et.xpath("Project/Site_Information/Remarks/text()")[0]
# Outside_Gage_Height = et.xpath("Project/Site_Information/Outside_Gage_Height/text()")[0]
# Water_Temperature = et.xpath("Project/Site_Information/Water_Temperature/text()")[0]
################################################################################################

# READ EXCEL FILE
# xlsfile = "Meritve.xlsx"
# filename = (r'C:\Users\nucic\Desktop\MERITVE\TestnoNoveMeritve')
xlsfile = (r'C:\Users\nucic\Desktop\Meritve.xlsx')
filename = (r'C:\Users\nucic\Desktop\MERITVE\NOVE')

#WORKBOOK
wb = openpyxl.load_workbook(xlsfile)
#GET FIRST SHEET
first_sheet = wb.get_sheet_names()[0]
ws = wb.active

# Get the last measurement
LastMeasNum = ws.max_row-1

# SET DIRECTORY LOCATION STATIC / DYNAMIC
# dir = os.path.dirname(__file__)
# filename = os.path.join(dir, 'Meritve')

files = os.listdir(filename)

######GET DATA FROM XML FILES ####################
def GetAndSaveDataFromXml(filee,LastMeasNum):

    # EXTRACT DATE FROM FILENAME!!!
    filename = filee[:-4]
    FilenameYear=filename[5:9]
    FilenameMonth = filename[10:12]
    FilenameDay = filename[13:15]
    # print (FilenameYear,FilenameMonth,FilenameDay)

    file = os.path.join(root, filee)
    et = etree.parse(file)

    ProcessedBy = et.xpath("Project/Site_Information/ProcessedBy/text()")[0]
    # print(ProcessedBy)
    Measurement_Date = et.xpath("Project/Site_Information/Measurement_Date/text()")[0]
    # EXTRACT DATE FROM MEASUREMENT DATE
    MeasYear = Measurement_Date[6:]
    MeasMonth= Measurement_Date[0:2]
    MeasDay= Measurement_Date[3:5]
    # print(MeasYear,MeasMonth,MeasDay)
    River_Name = et.xpath("Project/Site_Information/River_Name/text()")[0]
    Number = et.xpath("Project/Site_Information/Number/text()")[0]
    Name = et.xpath("Project/Site_Information/Name/text()")[0]
    Remarks = et.xpath("Project/Site_Information/Remarks/text()")
    ADCPTemp = et.xpath("Project/Site_Discharge/Discharge_Summary/None/Index_0/ADCPTemperature/text()")
    # print (ADCPTemp)
    # IF THERE IS NO REMARKS
    if len(Remarks) != 0:
        Remarks = et.xpath("Project/Site_Information/Remarks/text()")[0]
    else:
        Remarks = ""
    Outside_Gage_Height = format(float(et.xpath("Project/Site_Information/Outside_Gage_Height/text()")[0])*100,'.1f')
    Water_Temperature = format(float(et.xpath("Project/Site_Information/Water_Temperature/text()")[0]),'.1f')
    # print (Water_Temperature)

    # UGOTAVLJANJE RAZLIKE MED MERITVIJO IN IMENOM DATOTEK - ISKANJE NAPAK DATUMA IN ČASA
    PrimerjavaLeto = int(FilenameYear)-int(MeasYear)
    PrimerjavaMesec= int(FilenameMonth)-int(MeasMonth)
    PrimerjavaDan = int(FilenameDay)-int(MeasDay)
    print (Number,River_Name,Name,PrimerjavaLeto,PrimerjavaMesec,PrimerjavaDan)

# DOPIŠI PODATKE V EXCEL
    new_data_forTransferingToEXCEL = [LastMeasNum, filename,Measurement_Date, ProcessedBy, Number, River_Name, Name, Outside_Gage_Height, Water_Temperature, Remarks]
    ws.append(new_data_forTransferingToEXCEL)


# SEARCH FOR ALL MMT FILES AND ADD METADATA TO EXCEL
for root, dirs, files in os.walk(filename):
    for file in files:
        try:
            if (file.endswith(".mmt")):
                # print(file)
                LastMeasNum += 1
                GetAndSaveDataFromXml(file,LastMeasNum)
        except Exception as inst:
            print("Z datoteko: " + file + ", so nastale težave - preveri!!")
            print (inst)
            pass
    wb.save(xlsfile)